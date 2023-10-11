#规则

import log
import openpyxl
from google.protobuf.descriptor import FieldDescriptor

def get_field_type(cell_name,item):
    field_type = -1
    for field in item.DESCRIPTOR.fields:
        if field.name == cell_name:
            field_type = field.type
            break
    if field_type == -1:
        error_str  = "Error: excel cell name can not find in proto: "+cell_name
        log.debug(error_str)
    return field_type

def get_field_enum_type(cell_name,cell_value,item):
    for field in item.DESCRIPTOR.fields:
        if field.name == cell_name:
            return field.enum_type.name, field.enum_type.values_by_name[cell_value].number

def get_field_is_array(cell_name,item):
    is_array = False
    for field in item.DESCRIPTOR.fields:
        if field.name == cell_name:
            is_array = field.label == FieldDescriptor.LABEL_REPEATED
    return is_array

#解析excel单元格
def parse_cell(cell_name,cell_value,item):
    if len(str(cell_value).strip())<=0:
        error_str  = "Error: this cell is empty: "+cell_name
        log.debug(error_str)
        return None
    log_str = "parse_cell type:"
    field_type = get_field_type(cell_name,item)
    if field_type==FieldDescriptor.TYPE_INT32 or field_type==FieldDescriptor.TYPE_SINT32\
        or field_type==FieldDescriptor.TYPE_SFIXED32:
        log_str =log_str+cell_name+" "+ str(field_type)+" to int"
        log.debug(log_str)
        return int(cell_value)
    elif field_type==FieldDescriptor.TYPE_DOUBLE or field_type==FieldDescriptor.TYPE_FLOAT:
        log_str =log_str+cell_name+" "+ str(field_type)+" to float"
        log.debug(log_str)
        return float(cell_value)
    elif field_type==FieldDescriptor.TYPE_INT64 or field_type==FieldDescriptor.TYPE_UINT32\
        or field_type==FieldDescriptor.TYPE_UINT64 or field_type==FieldDescriptor.TYPE_SINT64\
        or field_type==FieldDescriptor.TYPE_FIXED32 or field_type==FieldDescriptor.TYPE_FIXED64\
        or field_type==FieldDescriptor.TYPE_SFIXED64:
        log_str =log_str+cell_name+" "+ str(field_type)+" to int"
        log.debug(log_str)
        return int(cell_value)
    elif field_type==FieldDescriptor.TYPE_BOOL:
        log_str =log_str+cell_name+" "+ str(field_type)+" to bool"
        log.debug(log_str)
        return bool(int(cell_value))
    elif field_type==FieldDescriptor.TYPE_STRING:
        log_str =log_str+cell_name+" "+ str(field_type)+" to str"
        log.debug(log_str)
        return str(cell_value)
    elif field_type==FieldDescriptor.TYPE_BYTES:
        log_str =log_str+cell_name+" "+ str(field_type)+" to bytes"
        log.debug(log_str)
        return bytes(cell_value)
    elif field_type==FieldDescriptor.TYPE_ENUM:
        enum_name, enum_value = get_field_enum_type(cell_name,cell_value,item)
        log_str =log_str+cell_name+" "+ str(field_type)+" to enum "+ enum_name
        log.debug(log_str)
        return enum_value
    log.debug("Error: parse_cell type:"+cell_name)
    return None

#解析excel单元格数组(以,分隔)
def parse_array_cell(cell_name,cell_value,item):
    log.debug("parse_array_cell "+cell_name)
    values = []
    cell_value_all = cell_value #只有一个元素
    cell_value_all = cell_value.split(',')
    for cell_value_single in cell_value_all:
        value = parse_cell(cell_name,cell_value_single,item)
        values.append(value)
    return values

def parse_field(cell_name,cell_value,item):    
    is_array = get_field_is_array(cell_name,item)
    log.debug("parse_field is_array: "+cell_name+" "+str(is_array))
    if is_array:
        values = parse_array_cell(cell_name,cell_value,item)
        field = getattr(item,cell_name)
        for value in values:
            field.append(value)
    else:
        value = parse_cell(cell_name,cell_value,item)
        setattr(item,cell_name,value)

def get_item_array(excel_path,item_array):
    log.debug("get_item_array "+excel_path)
    sheet = openpyxl.load_workbook(excel_path).worksheets[0]
    #第一行注释,第二行字段名,第三行以后配置数值
    row_count = sheet.max_row
    col_count = sheet.max_column
    #excel从1开始
    for row_index in range(row_count+1):
        if row_index>=3:
            item = item_array.item.add()
            for col_index in range(col_count+1):
                if col_index>=1:
                    cell_name = sheet.cell(2,col_index).value
                    cell_value = sheet.cell(row_index,col_index).value
                    cell_value = str(cell_value)
                    parse_field(cell_name,cell_value,item)
    return item_array